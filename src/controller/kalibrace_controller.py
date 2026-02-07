from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import threading
import time
import pandas as pd
import queue
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import math
import inspect
from IPython.display import display

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from model.Piezo_model import Piezo_model
    from model.MCU_model import MCU_model
    from controller.main_controller import MainController
    
class KalibraceController():
    def __init__(self, controller : "MainController", piezo_model, mcu_model):
        self.controller : MainController = controller
        self.piezo_model : Piezo_model = piezo_model #asi neni potreba - delat pres controller
        self.mcu_model : MCU_model = mcu_model #asi neni potreba - delat pres controller
        self.pracovni_slozka = None
        self.pracovni_soubor = None
        self.kalibracni_obrazek = None
        self.delka_kroku = None
        self.merena_vzdalenost = None
        self.pocet_kroku = None #pocet kroku pro mereni = pocet iteraci ve for smycce
        self.pocet_zaznamu = 10
        self.kalibrace = False
        
        self.vzorky = []
        self.teplota = []
        self.tlak = []
        self.vlhkost = []
        self.osvetleni = []
        self.poloha = 0
     
        #fronta pro vytvareni grafu:
        self.queue_graf = queue.Queue()
        
        #Zpracovani dat
        self.protokol ={"1" : "A/D převodník",
                        "2" : "Pulzy",
                        "3" : "Protokol"}
        
        #Strategie zpracovani a kalibrace
        self.strategie = {"Dopředná" : "Dopředná",
                          "Zpětná" : "Zpětná",
                          "Opakovatelnost" : "Opakovatelnost",
                          "Hystereze" : "Hystereze",
                          "Hystereze2" : "Hystereze2"}
        
    def protokol_kalibrace(self, protokol):
        print(f"[KalibraceController] vybrany protokol : {self.protokol[protokol]}")
        # print(f"[{self.__class__.__name__}] protokol: {self.controller.protokol_gui.vybrane_var.get()}") #stejna odpoved
        
    def nastavit_delku_kroku(self, krok):     
        try:
            self.delka_kroku = round(float(krok), 3)  
            print(f"[{self.__class__.__name__}] delka kroku je {self.delka_kroku} (um)")
        except ValueError:
            print("neplatne cislo!")
        
    def nastavit_delku_vzdalenost(self, vzdalenost):
        try:
            hodnota = float(vzdalenost) #prevod na cislo
        except ValueError:
            print(f"[{self.__class__.__name__}] CHYBA!! NEPLATNY VSTUP ({vzdalenost})")
            return
    
        hodnota = round(hodnota, 3)
        self.merena_vzdalenost = hodnota

        if hasattr(self, "entry_vzdalenost"):
            self.controller.kalibrace_gui.entry_vzdalenost.delete(0, 'end')
            self.controller.kalibrace_gui.entry_vzdalenost.insert(0, f"{hodnota:.3f}")

        print(f"[{self.__class__.__name__}] merena vzdalenost je {self.merena_vzdalenost:.3f} (µm)")
        
    def vybrat_pracovni_slozku(self):
        self.pracovni_slozka = filedialog.askdirectory(title="Pracovní složka")
        print(f"[{self.__class__.__name__}] složka {self.pracovni_slozka}")
        
        if self.pracovni_slozka:
            self.controller.kalibrace_gui.Entry_slozka_pracovni.config(state="normal")
            self.controller.kalibrace_gui.Entry_slozka_pracovni.delete(0, "end")
            self.controller.kalibrace_gui.Entry_slozka_pracovni.insert(0, f"{self.pracovni_slozka}")
            self.controller.kalibrace_gui.Entry_slozka_pracovni.config(state="readonly")
            
        else:
            self.controller.kalibrace_gui.Entry_slozka_pracovni.config(state="normal")
            self.controller.kalibrace_gui.Entry_slozka_pracovni.delete(0, "end")
            self.controller.kalibrace_gui.Entry_slozka_pracovni.insert(0, f"N/A")
            self.controller.kalibrace_gui.Entry_slozka_pracovni.config(state="readonly")
            
    def nastavit_vzorky(self, pocet):
        if int(pocet) > 150:
            print(f"[{self.__class__.__name__}] CHYBA!! : nastavený počet vzorků převyšuje číslo 150, nastaveno implicitne 150 !!")
            self.controller.kalibrace_gui.entry_pocet_vzorku.delete(0, 'end')
            self.controller.kalibrace_gui.entry_pocet_vzorku.insert(0, "150")
            self.pocet_zaznamu = 150
        else:
            self.pocet_zaznamu = pocet
            print(f"[{self.__class__.__name__}] nastavený počet vzorků je nastaven na {self.pocet_zaznamu}")
            
    def data_load(self, zpracovani_dat : str,  strategie_kalibrace: str, delka_kroku : float, merena_vzdalenost : float, pocet_vzorku : int, pocet_kroku : int):
        self.controller.zpracovani.kalibrace_df = pd.DataFrame([{
            "rozsah_mereni" : merena_vzdalenost,
            "rozliseni_mereni" : delka_kroku,
            "pocet_kroku" : pocet_kroku,
            "pocet_vzorku_na_krok" : pocet_vzorku,
            "zpracovani_dat" : zpracovani_dat,
            "strategie_kalibrace" : strategie_kalibrace
        }])
          
    #SMYCKA VLAKNO FUNKCE kalibrace s AD ZPETNA      
    def kalibrace_start_ad_zpetna(self):
        print(f"[{self.__class__.__name__}] [{inspect.currentframe().f_code.co_name}] SMYCKA VLAKNO FUNKCE!!!")      
        self.controller.blok_widgets(self.controller.root) #zablokovani widgetu
        
        #KONTROLA PROSTORU PRED KALIBRACI
        if self.controller.piezo_model.prostor == False:
                    print(f"[{self.__class__.__name__}] FUNKCE PREDCASNE UKONCENA !!! -- NESPLNENA PODMINKA. BYL PROVEDEN HOME?")
                    InfoMsg = f"[{self.__class__.__name__}] FUNKCE PREDCASNE UKONCENA !!! -- NESPLNENA PODMINKA. BYL PROVEDEN HOME?"
                    messagebox.showinfo("INFORMACE", InfoMsg)
                    return
                    
        #pracovni slozka:
        if self.pracovni_slozka is not None:
            cesta_csv = os.path.join(self.pracovni_slozka, f"temp.csv")
            cesta_xlsx = os.path.join(self.pracovni_slozka, f"temp.xlsx")
    
        #promazani DF datoveho modelu ve Zpracovani_model
        self.controller.zpracovani.kalibrace_df = pd.DataFrame()
        self.controller.zpracovani.df = pd.DataFrame()
        self.controller.zpracovani.summary_df = pd.DataFrame()
        
        #kalibrace jeste neskoncila
        self.controller.kalibrace_finish = False
        
        #vycisteni fronty
        with self.queue_graf.mutex:
            self.queue_graf.queue.clear() 
        
        #SMYCKA VLAKNO
        def kalibrace_start_inner():
            self.kalibrace = True #start kalibrace
            print(f"[{self.__class__.__name__}] VLAKNO KALIBRACE!")
            
            self.pocet_kroku = math.floor(self.merena_vzdalenost / self.delka_kroku)
            
            #zacatek vytvoreni docasneho souboru a zapis do nej
            df_header = pd.DataFrame(columns=["cas", "pozice", "napeti", "teplota", "tlak", "vlhkost", "osvetleni", "smer"])
            df_header.to_csv(cesta_csv, index=False, header=True)
            
            #cekani na najeti do referencni polohy zadane uzivatelem
            #zajede na pozici a ceka na dalsi ukoly
            while True:
                time.sleep(0.5)
                try:
                    if self.piezo_model.is_homed == True:
                        print(f"[{self.__class__.__name__}] HOME jiz provedeno NAJIZDENI DO REFERENCNI POLOHY !")
                        
                        #NAJETI ZPET DO REFERENCNI POLOHY
                        pohyb_x = self.piezo_model.x - self.piezo_model.x_ref
                        pohyb_y = self.piezo_model.y - self.piezo_model.y_ref
                        pohyb_z = self.piezo_model.z - self.piezo_model.z_ref
                        self.controller.M_C_send_msg_piezo(f"GT x{pohyb_x:.3f} y{pohyb_y:.3f} z{pohyb_z:.3f}")
                        #prodleva nez najede na pozici pro 
                        time.sleep(5)
                        break
                except Exception as e:
                    print(f"[{self.__class__.__name__}] CHYBA ({e})")
            #smycka pro sber dat a zapis do souboru 
            #doresit - vymenit while za for - a pocitat do max vzdalenosti -- inkrementace piezo dle vzdalenosti
            
            smer = "y"
            iterace = 0
            for _ in range(int(self.pocet_kroku) + 2):

                #pokud je kalibrace nekde ukoncena, tak preruseni iteraci
                if self.controller.piezo_model.prostor == False or self.kalibrace == False:
                    print(f"[{self.__class__.__name__}] FUNKCE PREDCASNE UKONCENA !!! -- NESPLNENA/Y PODMINKA. BYL PROVEDEN HOME?")
                    self.controller.kalibrace_finish = False
                    break       
                
                while self.mcu_model.lock_ad == False:
                        time.sleep(0.1) 

                if iterace > 0:
                    #zapsat mcu napeti a piezo polohu
                    self.vzorky = self.mcu_model.napeti_vzorky.copy()
                    self.teplota = self.mcu_model.teplota_vzorky.copy()
                    self.tlak = self.mcu_model.tlak_vzorky.copy()
                    self.vlhkost = self.mcu_model.vlhkost_vzorky.copy()
                    self.osvetleni = self.mcu_model.osvetleni_vzorky.copy()
                    self.poloha = round(float(self.piezo_model.y_ref) * 1, 3)
                    if abs(self.poloha) < 1e-9:
                        self.poloha = 0.0
                    cas = datetime.now().strftime("%H:%M:%S.%f")[:-3]  

                    df = pd.DataFrame({
                        "cas": [cas]*len(self.vzorky),
                        "pozice": [f"{self.poloha:.3f}"]*len(self.vzorky),
                        "napeti": self.vzorky,
                        "teplota" : self.teplota,
                        "tlak": self.tlak,
                        "vlhkost": self.vlhkost,
                        "osvetleni" : self.osvetleni,
                        "smer" : [smer]*len(self.vzorky)
                    })  
                    self.controller.zpracovani.df = pd.concat([self.controller.zpracovani.df, df], ignore_index = True)
                    
                    #ulozeni surovych dat do csv
                    df.to_csv(cesta_csv, mode='a', header=False, index=False)

                    #zapis statistickych hodnot do dat a poslani do Zpracovani_modelu tridy
                    summary_df = pd.DataFrame([{
                        "cas": cas,
                        "pozice": self.poloha,
                        "napeti_prumer": df["napeti"].mean(),
                        "napeti_std": df["napeti"].std(),
                        "teplota_prumer": df["teplota"].mean(),
                        "tlak_prumer": df["tlak"].mean(),
                        "vlhkost_prumer": df["vlhkost"].mean(),
                        "osvetleni_prumer" : df["osvetleni"].mean(),
                        "smer" : smer
                    }])
                    self.controller.zpracovani.summary_df = pd.concat([self.controller.zpracovani.summary_df, summary_df], ignore_index=True)

                    #pridani jednotlivych polozek vzorku do fronty
                    for n, t in zip(self.vzorky, self.teplota):
                        self.queue_graf.put({
                            "cas" : cas,
                            "pozice" : self.poloha,
                            "napeti" : n,
                            "teplota" : t,
                            "smer" : smer
                        })

                    print(f"[{self.__class__.__name__}] zápis dat do fronty!")

                    if iterace <= int(self.pocet_kroku):    
                            self.controller.M_C_nastav_pohyb_piezo(float(self.delka_kroku))
                            self.controller.M_C_pohyb_piezo("y")
                            while not self.controller.lock_pohyb:
                                time.sleep(0.1)     

                #nastavit pozici - prvi iterace nulove posunuti

                self.mcu_model.lock_ad = False

                if iterace <= int(self.pocet_kroku):
                    print(f"[{self.__class__.__name__}] iterace")
                    print(f"[{self.__class__.__name__}] CTENI NAPETI !!")
                    self.mcu_model.precist_AD(int(self.pocet_zaznamu))

                iterace += 1

                #pokud vsechny pozice bylo zmereno, tak uspesna kalibrace
                if iterace == (self.pocet_kroku):
                    self.controller.kalibrace_finish = True
                            
            #precteni celeho CSV a nahrani do XLSX
            df_csv = pd.read_csv(cesta_csv)
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Data"
                #hlavicka
                hlavicka = ["Čas (hh:mm:ss)", "Pozice (µm)", "Napětí (V)", "Teplota (°C)", "Tlak (Pa)", "Vlhkost (%)", "Osvetleni (lux)"]
                bold_font = Font(bold=True)
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                #zapsat hlavicku
                for col_num, text in enumerate(hlavicka, start=1):
                    cell = ws.cell(row=1, column=col_num, value=text)
                    cell.font = bold_font
                #zapsat data
                for row_num, row_data in enumerate(df_csv.values, start=2):
                    for col_num, value in enumerate(row_data, start=1):
                        ws.cell(row=row_num, column=col_num, value=value)
                #ulozit
                wb.save(cesta_xlsx)
                print(f"[{self.__class__.__name__}] EXCEL VZORKY VYTVORENY, EXCEL (temp soubor) VYTVOREN") 
                self.controller.kalibrace_finish = True
                self.data_load(zpracovani_dat="A/D", strategie_kalibrace="zpětná", delka_kroku=self.delka_kroku, merena_vzdalenost=self.merena_vzdalenost ,pocet_vzorku=self.pocet_zaznamu, pocet_kroku=self.pocet_kroku)    
                if self.controller.kalibrace_finish == False:
                    InfoMsg = f"Data byly uloženy ale kalibrace neproběhla celá"
                else:
                    InfoMsg = f"Kalibrace proběhla úspěšně"
                messagebox.showinfo("Kalibrace", InfoMsg)
            except Exception as e:
                self.kalibrace = False
                InfoMsg = f"CHYBA\nEXCEL (temp soubor) VZORKY NEVYTVORENY KONEC APLIKACE !!CHYBA!!\npravdepodobne otevreny soubor temp:\n{e}!!"
                messagebox.showinfo("Chyba", InfoMsg)
                print(f"[{self.__class__.__name__}] EXCEL VZORKY NEVYTVORENY --CHYBA!! {e}")   

            print(f"[{self.__class__.__name__}] sběr dat hotový")  
            self.kalibrace = False #konec kalibrace
            self.mcu_model.lock_ad = True #zase odemknout pro pripad dalsiho mereni
            
            #debug vypis dat rychly
            display(self.controller.zpracovani.df)
            display(self.controller.zpracovani.summary_df)
            
        self.t1 = threading.Thread(target=kalibrace_start_inner, daemon=True)
        self.t1.start()
    
    #SMYCKA VLAKNO FUNKCE kalibrace s AD hystereze     
    def kalibrace_start_ad_hystereze(self):
        print(f"[{self.__class__.__name__}] [{inspect.currentframe().f_code.co_name}] SMYCKA VLAKNO FUNKCE!!!")      
        self.controller.blok_widgets(self.controller.root) #zablokovani widgetu
        
        #KONTROLA PROSTORU PRED KALIBRACI
        if self.controller.piezo_model.prostor == False:
                    print(f"[{self.__class__.__name__}] FUNKCE PREDCASNE UKONCENA !!! -- NESPLNENA PODMINKA. BYL PROVEDEN HOME?")
                    InfoMsg = f"[{self.__class__.__name__}] FUNKCE PREDCASNE UKONCENA !!! -- NESPLNENA PODMINKA. BYL PROVEDEN HOME?"
                    messagebox.showinfo("INFORMACE", InfoMsg)
                    return
                    
        #pracovni slozka:
        if self.pracovni_slozka is not None:
            cesta_csv = os.path.join(self.pracovni_slozka, f"temp.csv")
            cesta_xlsx = os.path.join(self.pracovni_slozka, f"temp.xlsx")
    
        #promazani DF datoveho modelu ve Zpracovani_model
        self.controller.zpracovani.kalibrace_df = pd.DataFrame()
        self.controller.zpracovani.df = pd.DataFrame()
        self.controller.zpracovani.summary_df = pd.DataFrame()
        
        #kalibrace jeste neskoncila
        self.controller.kalibrace_finish = False
        
        #vycisteni fronty
        with self.queue_graf.mutex:
            self.queue_graf.queue.clear() 
        
        #SMYCKA VLAKNO
        def kalibrace_start_inner():
            self.kalibrace = True #start kalibrace
            print(f"[{self.__class__.__name__}] VLAKNO KALIBRACE!")
            
            self.pocet_kroku = math.floor(self.merena_vzdalenost / self.delka_kroku)
            
            #zacatek vytvoreni docasneho souboru a zapis do nej
            df_header = pd.DataFrame(columns=["cas", "pozice", "napeti", "teplota", "tlak", "vlhkost", "osvetleni", "smer"])
            df_header.to_csv(cesta_csv, index=False, header=True)
            
            #cekani na najeti do referencni polohy zadane uzivatelem
            #zajede na pozici a ceka na dalsi ukoly
            while True:
                time.sleep(0.5)
                try:
                    if self.piezo_model.is_homed == True:
                        print(f"[{self.__class__.__name__}] HOME jiz provedeno NAJIZDENI DO REFERENCNI POLOHY !")
                        
                        #NAJETI ZPET DO REFERENCNI POLOHY
                        pohyb_x = self.piezo_model.x - self.piezo_model.x_ref
                        pohyb_y = self.piezo_model.y - self.piezo_model.y_ref
                        pohyb_z = self.piezo_model.z - self.piezo_model.z_ref
                        self.controller.M_C_send_msg_piezo(f"GT x{pohyb_x:.3f} y{pohyb_y:.3f} z{pohyb_z:.3f}")
                        #prodleva nez najede na pozici pro 
                        time.sleep(5)
                        break
                except Exception as e:
                    print(f"[{self.__class__.__name__}] CHYBA ({e})")
            #smycka pro sber dat a zapis do souboru 
            #doresit - vymenit while za for - a pocitat do max vzdalenosti -- inkrementace piezo dle vzdalenosti
            
            #SMER PRIBLIZOVANI
            smer = "y"
            iterace = 0
            for _ in range(int(self.pocet_kroku) + 2):

                #pokud je kalibrace nekde ukoncena, tak preruseni iteraci
                if self.controller.piezo_model.prostor == False or self.kalibrace == False:
                    print(f"[{self.__class__.__name__}] FUNKCE PREDCASNE UKONCENA !!! -- NESPLNENA/Y PODMINKA. BYL PROVEDEN HOME?")
                    self.controller.kalibrace_finish = False
                    break       
                
                while self.mcu_model.lock_ad == False:
                        time.sleep(0.1) 

                if iterace > 0:
                    #zapsat mcu napeti a piezo polohu
                    self.vzorky = self.mcu_model.napeti_vzorky.copy()
                    self.teplota = self.mcu_model.teplota_vzorky.copy()
                    self.tlak = self.mcu_model.tlak_vzorky.copy()
                    self.vlhkost = self.mcu_model.vlhkost_vzorky.copy()
                    self.osvetleni = self.mcu_model.osvetleni_vzorky.copy()
                    self.poloha = round(float(self.piezo_model.y_ref) * 1, 3)
                    if abs(self.poloha) < 1e-9:
                        self.poloha = 0.0
                    cas = datetime.now().strftime("%H:%M:%S.%f")[:-3]  

                    df = pd.DataFrame({
                        "cas": [cas]*len(self.vzorky),
                        "pozice": [f"{self.poloha:.3f}"]*len(self.vzorky),
                        "napeti": self.vzorky,
                        "teplota" : self.teplota,
                        "tlak": self.tlak,
                        "vlhkost": self.vlhkost,
                        "osvetleni" : self.osvetleni,
                        "smer" : [smer]*len(self.vzorky)
                    })  
                    self.controller.zpracovani.df = pd.concat([self.controller.zpracovani.df, df], ignore_index = True)
                    
                    #ulozeni surovych dat do csv
                    df.to_csv(cesta_csv, mode='a', header=False, index=False)

                    #zapis statistickych hodnot do dat a poslani do Zpracovani_modelu tridy
                    summary_df = pd.DataFrame([{
                        "cas": cas,
                        "pozice": self.poloha,
                        "napeti_prumer": df["napeti"].mean(),
                        "napeti_std": df["napeti"].std(),
                        "teplota_prumer": df["teplota"].mean(),
                        "tlak_prumer": df["tlak"].mean(),
                        "vlhkost_prumer": df["vlhkost"].mean(),
                        "osvetleni_prumer" : df["osvetleni"].mean(),
                        "smer" : smer
                    }])
                    self.controller.zpracovani.summary_df = pd.concat([self.controller.zpracovani.summary_df, summary_df], ignore_index=True)

                    #pridani jednotlivych polozek vzorku do fronty
                    for n, t in zip(self.vzorky, self.teplota):
                        self.queue_graf.put({
                            "cas" : cas,
                            "pozice" : self.poloha,
                            "napeti" : n,
                            "teplota" : t,
                            "smer" : smer
                        })

                    print(f"[{self.__class__.__name__}] zápis dat do fronty!")

                    if iterace <= int(self.pocet_kroku):    
                            self.controller.M_C_nastav_pohyb_piezo(float(self.delka_kroku))
                            self.controller.M_C_pohyb_piezo(smer)
                            while not self.controller.lock_pohyb:
                                time.sleep(0.1)     
                                
                self.mcu_model.lock_ad = False

                if iterace <= int(self.pocet_kroku):
                    print(f"[{self.__class__.__name__}] iterace")
                    print(f"[{self.__class__.__name__}] CTENI NAPETI !!")
                    self.mcu_model.precist_AD(int(self.pocet_zaznamu))

                iterace += 1

            #SMER oddalovani
            self.mcu_model.lock_ad = True #odemknuti ad
            smer = "y-"
            iterace = 0
            for _ in range(int(self.pocet_kroku) + 2):

                #pokud je kalibrace nekde ukoncena, tak preruseni iteraci
                if self.controller.piezo_model.prostor == False or self.kalibrace == False:
                    print(f"[{self.__class__.__name__}] FUNKCE PREDCASNE UKONCENA !!! -- NESPLNENA/Y PODMINKA. BYL PROVEDEN HOME?")
                    self.controller.kalibrace_finish = False
                    break       
                
                while self.mcu_model.lock_ad == False:
                        time.sleep(0.1) 

                if iterace > 0:
                    #zapsat mcu napeti a piezo polohu
                    self.vzorky = self.mcu_model.napeti_vzorky.copy()
                    self.teplota = self.mcu_model.teplota_vzorky.copy()
                    self.tlak = self.mcu_model.tlak_vzorky.copy()
                    self.vlhkost = self.mcu_model.vlhkost_vzorky.copy()
                    self.osvetleni = self.mcu_model.osvetleni_vzorky.copy()
                    self.poloha = round(float(self.piezo_model.y_ref) * 1, 3)
                    if abs(self.poloha) < 1e-9:
                        self.poloha = 0.0
                    cas = datetime.now().strftime("%H:%M:%S.%f")[:-3]  

                    df = pd.DataFrame({
                        "cas": [cas]*len(self.vzorky),
                        "pozice": [f"{self.poloha:.3f}"]*len(self.vzorky),
                        "napeti": self.vzorky,
                        "teplota" : self.teplota,
                        "tlak": self.tlak,
                        "vlhkost": self.vlhkost,
                        "osvetleni" : self.osvetleni,
                        "smer" : [smer]*len(self.vzorky)
                    })  
                    self.controller.zpracovani.df = pd.concat([self.controller.zpracovani.df, df], ignore_index = True)
                    
                    #ulozeni surovych dat do csv
                    df.to_csv(cesta_csv, mode='a', header=False, index=False)

                    #zapis statistickych hodnot do dat a poslani do Zpracovani_modelu tridy
                    summary_df = pd.DataFrame([{
                        "cas": cas,
                        "pozice": self.poloha,
                        "napeti_prumer": df["napeti"].mean(),
                        "napeti_std": df["napeti"].std(),
                        "teplota_prumer": df["teplota"].mean(),
                        "tlak_prumer": df["tlak"].mean(),
                        "vlhkost_prumer": df["vlhkost"].mean(),
                        "osvetleni_prumer" : df["osvetleni"].mean(),
                        "smer" : smer
                    }])
                    self.controller.zpracovani.summary_df = pd.concat([self.controller.zpracovani.summary_df, summary_df], ignore_index=True)

                    #pridani jednotlivych polozek vzorku do fronty
                    for n, t in zip(self.vzorky, self.teplota):
                        self.queue_graf.put({
                            "cas" : cas,
                            "pozice" : self.poloha,
                            "napeti" : n,
                            "teplota" : t,
                            "smer" : smer
                        })

                    print(f"[{self.__class__.__name__}] zápis dat do fronty!")

                    if iterace <= int(self.pocet_kroku):    
                            self.controller.M_C_nastav_pohyb_piezo(float(self.delka_kroku))
                            self.controller.M_C_pohyb_piezo(smer)
                            while not self.controller.lock_pohyb:
                                time.sleep(0.1)     
                            
                self.mcu_model.lock_ad = False

                if iterace <= int(self.pocet_kroku):
                    print(f"[{self.__class__.__name__}] iterace")
                    print(f"[{self.__class__.__name__}] CTENI NAPETI !!")
                    self.mcu_model.precist_AD(int(self.pocet_zaznamu))

                iterace += 1

                #pokud vsechny pozice bylo zmereno, tak uspesna kalibrace
                if iterace == (self.pocet_kroku):
                    self.controller.kalibrace_finish = True
                
                
            #precteni celeho CSV a nahrani do XLSX
            df_csv = pd.read_csv(cesta_csv)
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Data"
                #hlavicka
                hlavicka = ["Čas (hh:mm:ss)", "Pozice (µm)", "Napětí (V)", "Teplota (°C)", "Tlak (Pa)", "Vlhkost (%)", "Osvetleni (lux)"]
                bold_font = Font(bold=True)
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                #zapsat hlavicku
                for col_num, text in enumerate(hlavicka, start=1):
                    cell = ws.cell(row=1, column=col_num, value=text)
                    cell.font = bold_font
                #zapsat data
                for row_num, row_data in enumerate(df_csv.values, start=2):
                    for col_num, value in enumerate(row_data, start=1):
                        ws.cell(row=row_num, column=col_num, value=value)
                #ulozit
                wb.save(cesta_xlsx)
                print(f"[{self.__class__.__name__}] EXCEL VZORKY VYTVORENY, EXCEL (temp soubor) VYTVOREN") 
                self.controller.kalibrace_finish = True
                self.data_load(zpracovani_dat="A/D", strategie_kalibrace="zpětná", delka_kroku=self.delka_kroku, merena_vzdalenost=self.merena_vzdalenost ,pocet_vzorku=self.pocet_zaznamu, pocet_kroku=self.pocet_kroku)    
                if self.controller.kalibrace_finish == False:
                    InfoMsg = f"Data byly uloženy ale kalibrace neproběhla celá"
                else:
                    InfoMsg = f"Kalibrace proběhla úspěšně"
                messagebox.showinfo("Kalibrace", InfoMsg)
            except Exception as e:
                self.kalibrace = False
                InfoMsg = f"CHYBA\nEXCEL (temp soubor) VZORKY NEVYTVORENY KONEC APLIKACE !!CHYBA!!\npravdepodobne otevreny soubor temp:\n{e}!!"
                messagebox.showinfo("Chyba", InfoMsg)
                print(f"[{self.__class__.__name__}] EXCEL VZORKY NEVYTVORENY --CHYBA!! {e}")   

            print(f"[{self.__class__.__name__}] sběr dat hotový")  
            self.kalibrace = False #konec kalibrace
            self.mcu_model.lock_ad = True #zase odemknout pro pripad dalsiho mereni
            
            #debug vypis dat rychly
            display(self.controller.zpracovani.df)
            display(self.controller.zpracovani.summary_df)
            
        self.t1 = threading.Thread(target=kalibrace_start_inner, daemon=True)
        self.t1.start()
       
    #SMYCKA VLAKNO FUNKCE kalibrace s PULZY DOPREDNA
    """
    SMYCKA PRO PULZY DOPREDNA
    """            
    def kalibrace_start_pulzy_dopredna(self):
        print(f"[{self.__class__.__name__}] [{inspect.currentframe().f_code.co_name}] SMYCKA VLAKNO FUNKCE!!!")
        #self.controller.M_C_Index()
        self.controller.blok_widgets(self.controller.root) #zablokovani widgetu - M_C_Index neco odblokuje
        #pracovni slozka:
        if self.pracovni_slozka is not None:
            cesta_csv = os.path.join(self.pracovni_slozka, f"temp.csv")
            cesta_xlsx = os.path.join(self.pracovni_slozka, f"temp.xlsx")
         
        #promazani DF datoveho modelu ve Zpracovani_model
        self.controller.zpracovani.kalibrace_df = pd.DataFrame()
        self.controller.zpracovani.df = pd.DataFrame()
        self.controller.zpracovani.summary_df = pd.DataFrame()
        
        #kalibrace jeste neskoncila
        self.controller.kalibrace_finish = False
        
        #vycisteni fronty
        with self.queue_graf.mutex:
            self.queue_graf.queue.clear() 
                
        #SMYCKA VLAKNO
        def kalibrace_start_inner():
            self.kalibrace = True #start kalibrace
            time.sleep(5) #delay kvuli index pozici
            print(f"[{self.__class__.__name__}] VLAKNO KALIBRACE!")
            self.pocet_kroku = math.floor(self.merena_vzdalenost / self.delka_kroku)
            
            
            #zacatek vytvoreni docasneho souboru a zapis do nej
            df_header = pd.DataFrame(columns=["cas", "pozice", "frekvence", "teplota", "tlak", "vlhkost","osvetleni", "smer"])
            df_header.to_csv(cesta_csv, index=False, header=True)
          
            #cekani na vychozi pozici kalibrace
            #zajede na pozici a ceka na dalsi ukoly
            while True:
                time.sleep(0.5)
                try:
                    if self.piezo_model.is_homed == True:
                        print(f"[{self.__class__.__name__}] HOMED!")
                        #self.controller.M_C_send_msg_piezo("GT x0 y10000 z-5000") #pozice - max v Y
                        self.controller.M_C_send_msg_piezo("GT x0 y10000")
                        time.sleep(5) #CAS NEZ DOJEDE z home NA POZICI UDANE V self.controller.M_C_send_msg_piezo("GT x0 y10000 z5000") #TOTO MOZNA OPRAVIT TODO
                        self.controller.M_C_nastav_referenci() #jakmile stoji, tak se zreferuje pozice
                        time.sleep(0.5)
                        break
                except Exception as e:
                    print(f"[{self.__class__.__name__}] CHYBA ({e})")
                    
            #smycka pro sber dat a zapis do souboru 
            #doresit - vymenit while za for - a pocitat do max vzdalenosti -- inkrementace piezo dle vzdalenosti
            #SMER ODDALOVANI
            smer = "y-"
            iterace = 0
            for _ in range(int(self.pocet_kroku) + 2):
                
                #pokud je kalibrace nekde ukoncena, tak preruseni iteraci
                if self.controller.piezo_model.prostor == False or self.kalibrace == False:
                    print(f"[{self.__class__.__name__}] FUNKCE PREDCASNE UKONCENA !!! -- NESPLNENA/Y PODMINKA")
                    self.controller.kalibrace_finish = False
                    break
                
                while self.mcu_model.lock_frekvence == False:
                    time.sleep(0.1) 
                    
                if iterace > 0:
                    #zapsat mcu frekvence a piezo polohu
                    self.vzorky = self.mcu_model.frekvence_vzorky.copy()
                    self.teplota = self.mcu_model.teplota_vzorky.copy()
                    self.tlak = self.mcu_model.tlak_vzorky.copy()
                    self.vlhkost = self.mcu_model.vlhkost_vzorky.copy()
                    self.osvetleni = self.mcu_model.osvetleni_vzorky.copy()
                    self.poloha = round(float(self.piezo_model.y_ref) * -1, 3)
                    if abs(self.poloha) < 1e-9:
                        self.poloha = 0.000
                    cas = datetime.now().strftime("%H:%M:%S.%f")[:-3]
                                
                    df = pd.DataFrame({
                            "cas": [cas]*len(self.vzorky),
                            "pozice": [f"{self.poloha:.3f}"]*len(self.vzorky),
                            "frekvence": self.vzorky,
                            "teplota" : self.teplota,
                            "tlak": self.tlak,
                            "vlhkost": self.vlhkost,
                            "osvetleni" : self.osvetleni,
                            "smer" : [smer]*len(self.vzorky)
                        })
                    self.controller.zpracovani.df = pd.concat([self.controller.zpracovani.df, df], ignore_index = True)
                    
                    #ulozeni surovych dat do csv
                    df.to_csv(cesta_csv, mode='a', header=False, index=False)
                    
                    
                    #zapis statistickych hodnot do dat a poslani do Zpracovani_modelu tridy
                    summary_df = pd.DataFrame([{
                        "cas": cas,
                        "pozice": self.poloha,
                        "frekvence_prumer": df["frekvence"].mean(),
                        "frekvence_std": df["frekvence"].std(),
                        "teplota_prumer": df["teplota"].mean(),
                        "tlak_prumer": df["tlak"].mean(),
                        "vlhkost_prumer": df["vlhkost"].mean(),
                        "osvetleni_prumer" : df["osvetleni"].mean(),
                        "smer" : smer
                    }])
                    self.controller.zpracovani.summary_df = pd.concat([self.controller.zpracovani.summary_df, summary_df], ignore_index=True)
                                                     
                    #pridani jednotlivych polozek vzorku do fronty
                    for f, t in zip(self.vzorky, self.teplota):
                        self.queue_graf.put({
                            "cas" : cas,
                            "pozice" : self.poloha,
                            "frekvence" : f,
                            "teplota" : t,
                            "smer" : smer
                        })
                    
                    print(f"[{self.__class__.__name__}] zápis dat do fronty!") 
  
                    if iterace < int(self.pocet_kroku) + 1:    
                        self.controller.M_C_nastav_pohyb_piezo(float(self.delka_kroku))
                        self.controller.M_C_pohyb_piezo("y-")
                        while not self.controller.lock_pohyb:
                            time.sleep(0.1)               
                            
                #nastavit pozici - prvi iterace nulove posunuti
                #time sleep dle rychlosti    
                self.mcu_model.lock_frekvence = False
                
                if iterace <= int(self.pocet_kroku):
                    print(f"[{self.__class__.__name__}] iterace")
                    print(f"[{self.__class__.__name__}] CTENI FREKVENCE !!")
                    self.mcu_model.precist_frekvenci(int(self.pocet_zaznamu))
                
                iterace += 1
                
                #pokud vsechny pozice bylo zmereno, tak uspesna kalibrace
                if iterace == (self.pocet_kroku):
                    self.controller.kalibrace_finish = True
                  
            #precteni celeho CSV a nahrani do XLSX
            df_csv = pd.read_csv(cesta_csv)
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Data"
                #hlavicka
                hlavicka = ["Čas (hh:mm:ss)", "Pozice (µm)", "Frekvence (Hz)", "Teplota (°C)", "Tlak (Pa)", "Vlhkost (%)", "Osvetleni (lux)", "Smer (y-/y+)"]
                bold_font = Font(bold=True)
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                #zapsat hlavicku
                for col_num, text in enumerate(hlavicka, start=1):
                    cell = ws.cell(row=1, column=col_num, value=text)
                    cell.font = bold_font
                #zapsat data
                for row_num, row_data in enumerate(df_csv.values, start=2):
                    for col_num, value in enumerate(row_data, start=1):
                        ws.cell(row=row_num, column=col_num, value=value)
                #ulozit
                wb.save(cesta_xlsx)
                print(f"[{self.__class__.__name__}] EXCEL VZORKY VYTVORENY, EXCEL (temp soubor) VYTVOREN")
                self.data_load(zpracovani_dat="pulzy", strategie_kalibrace="dopředná", delka_kroku=self.delka_kroku,merena_vzdalenost=self.merena_vzdalenost ,pocet_vzorku=self.pocet_zaznamu, pocet_kroku=self.pocet_kroku) 
                if self.controller.kalibrace_finish == False:
                    InfoMsg = f"Data byly uloženy ale kalibrace neproběhla celá"
                else:
                    InfoMsg = f"Kalibrace proběhla úspěšně"
                messagebox.showinfo("Kalibrace", InfoMsg)
            except Exception as e:
                self.kalibrace = False
                InfoMsg = f"CHYBA\nEXCEL (temp soubor) VZORKY NEVYTVORENY KONEC APLIKACE !!CHYBA!!\npravdepodobne otevreny soubor temp:\n{e}!!"
                messagebox.showinfo("Chyba", InfoMsg)
                print(f"[{self.__class__.__name__}] EXCEL VZORKY NEVYTVORENY --CHYBA!! {e}")   
                  
            print(f"[{self.__class__.__name__}] sběr dat hotový")  
            self.kalibrace = False #konec kalibrace
            self.mcu_model.lock_frekvence = True #zase odemknout pro pripad dalsiho mereni
            
            #debug vypis dat rychly
            display(self.controller.zpracovani.df)
            display(self.controller.zpracovani.summary_df)
                        
        self.t1 = threading.Thread(target=kalibrace_start_inner, daemon=True)
        self.t1.start()
        
    #SMYCKA VLAKNO FUNKCE kalibrace s HYSTEREZE DOPREDNA      
    """
    SMYCKA PRO PULZY HYSTEREZE
    """    
    def kalibrace_start_pulzy_hystereze(self):
        print(f"[{self.__class__.__name__}] [{inspect.currentframe().f_code.co_name}] SMYCKA VLAKNO FUNKCE!!!")
        if self.controller.M_C_kalibracni_poloha_piezo() == 0:
            return
        time.sleep(3) #dokud zajede na konec
        #self.controller.M_C_Index()
        # while self.controller.piezo_model.piezo_serial.lock == False:
            # time.sleep(0.5)
            
        self.controller.blok_widgets(self.controller.root) #zablokovani widgetu - M_C_Index neco odblokuje
        #pracovni slozka:
        if self.pracovni_slozka is not None:
            cesta_csv = os.path.join(self.pracovni_slozka, f"temp.csv")
            cesta_xlsx = os.path.join(self.pracovni_slozka, f"temp.xlsx")
         
        #promazani DF datoveho modelu ve Zpracovani_model
        self.controller.zpracovani.kalibrace_df = pd.DataFrame()
        self.controller.zpracovani.df = pd.DataFrame()
        self.controller.zpracovani.summary_df = pd.DataFrame()
        
        #kalibrace jeste neskoncila
        self.controller.kalibrace_finish = False
        
        #vycisteni fronty
        with self.queue_graf.mutex:
            self.queue_graf.queue.clear() 
                
        #SMYCKA VLAKNO
        def kalibrace_start_inner():
            self.kalibrace = True #start kalibrace
            time.sleep(1) #delay kvuli index pozici
            print(f"[{self.__class__.__name__}] VLAKNO KALIBRACE!")
            self.pocet_kroku = math.floor(self.merena_vzdalenost / self.delka_kroku)
            
            
            #zacatek vytvoreni docasneho souboru a zapis do nej
            df_header = pd.DataFrame(columns=["cas", "pozice", "frekvence", "teplota", "tlak", "vlhkost","osvetleni", "smer"])
            df_header.to_csv(cesta_csv, index=False, header=True)
          
            #cekani na vychozi pozici kalibrace
            #zajede na pozici a ceka na dalsi ukoly
            while True:
                time.sleep(0.5)
                try:
                    if self.piezo_model.is_homed == True:
                        print(f"[{self.__class__.__name__}] HOMED!")
                        #self.controller.M_C_send_msg_piezo("GT x0 y10000 z-5000") #pozice - max v Y
                        # self.controller.M_C_send_msg_piezo("GT x0 y10000")
                        #time.sleep(2) #CAS NEZ DOJEDE z home NA POZICI UDANE V self.controller.M_C_send_msg_piezo("GT x0 y10000 z5000") #TOTO MOZNA OPRAVIT TODO
                        # while self.controller.piezo_model.piezo_serial.lock == False:
                            # time.sleep(0.5)
                        self.controller.M_C_nastav_referenci() #jakmile stoji, tak se zreferuje pozice
                        time.sleep(0.5)
                        break
                except Exception as e:
                    print(f"[{self.__class__.__name__}] CHYBA ({e})")
            #smycka pro sber dat a zapis do souboru 
            #doresit - vymenit while za for - a pocitat do max vzdalenosti -- inkrementace piezo dle vzdalenosti
            
            #SMER ODDALOVANI
            smer = "y-"
            iterace = 0
            for _ in range(int(self.pocet_kroku) + 2):
                
                #pokud je kalibrace nekde ukoncena, tak preruseni iteraci
                if self.controller.piezo_model.prostor == False or self.kalibrace == False:
                    print(f"[{self.__class__.__name__}] FUNKCE PREDCASNE UKONCENA !!! -- NESPLNENA/Y PODMINKA")
                    self.controller.kalibrace_finish = False
                    break
                
                while self.mcu_model.lock_frekvence == False:
                    time.sleep(0.1) 
                    
                if iterace > 0:
                    #zapsat mcu frekvence a piezo polohu
                    self.vzorky = self.mcu_model.frekvence_vzorky.copy()
                    self.teplota = self.mcu_model.teplota_vzorky.copy()
                    self.tlak = self.mcu_model.tlak_vzorky.copy()
                    self.vlhkost = self.mcu_model.vlhkost_vzorky.copy()
                    self.osvetleni = self.mcu_model.osvetleni_vzorky.copy()
                    self.poloha = round(float(self.piezo_model.y_ref) * -1, 3)
                    if abs(self.poloha) < 1e-9:
                        self.poloha = 0.000
                    cas = datetime.now().strftime("%H:%M:%S.%f")[:-3]
                                                      
                    df = pd.DataFrame({
                            "cas": [cas]*len(self.vzorky),
                            "pozice": [f"{self.poloha:.3f}"]*len(self.vzorky),
                            "frekvence": self.vzorky,
                            "teplota" : self.teplota,
                            "tlak": self.tlak,
                            "vlhkost": self.vlhkost,
                            "osvetleni" : self.osvetleni,
                            "smer" : [smer]*len(self.vzorky)
                        })
                    self.controller.zpracovani.df = pd.concat([self.controller.zpracovani.df, df], ignore_index = True)
                    
                    #ulozeni surovych dat do csv
                    df.to_csv(cesta_csv, mode='a', header=False, index=False)
                    
                    
                    #zapis statistickych hodnot do dat a poslani do Zpracovani_modelu tridy
                    summary_df = pd.DataFrame([{
                        "cas": cas,
                        "pozice": self.poloha,
                        "frekvence_prumer": df["frekvence"].mean(),
                        "frekvence_std": df["frekvence"].std(),
                        "teplota_prumer": df["teplota"].mean(),
                        "tlak_prumer": df["tlak"].mean(),
                        "vlhkost_prumer": df["vlhkost"].mean(),
                        "osvetleni_prumer" : df["osvetleni"].mean(),
                        "smer" : smer
                    }])
                    self.controller.zpracovani.summary_df = pd.concat([self.controller.zpracovani.summary_df, summary_df], ignore_index=True)
                                                                    
                    #pridani jednotlivych polozek vzorku do fronty
                    for f, t in zip(self.vzorky, self.teplota):
                        self.queue_graf.put({
                            "cas" : cas,
                            "pozice" : self.poloha,
                            "frekvence" : f,
                            "teplota" : t,
                            "smer" : smer
                        })
                    
                    print(f"[{self.__class__.__name__}] zápis dat do fronty!") 
  
                    if iterace <= int(self.pocet_kroku):    
                        self.controller.M_C_nastav_pohyb_piezo(float(self.delka_kroku))
                        self.controller.M_C_pohyb_piezo(smer)
                        while not self.controller.lock_pohyb:
                            time.sleep(0.1)               
                            
                #nastavit pozici - prvi iterace nulove posunuti
                #time sleep dle rychlosti    
                self.mcu_model.lock_frekvence = False
                
                if iterace <= int(self.pocet_kroku):
                    print(f"[{self.__class__.__name__}] iterace")
                    print(f"[{self.__class__.__name__}] CTENI FREKVENCE !!")
                    self.mcu_model.precist_frekvenci(int(self.pocet_zaznamu))
                  
                iterace += 1
                
            #SMER PRIBLIZOVANI
            self.mcu_model.lock_frekvence = True #odemknuti frekvence
            iterace = 0
            smer = "y"
            for _ in range(int(self.pocet_kroku) + 2):
                
                #pokud je kalibrace nekde ukoncena, tak preruseni iteraci
                if self.controller.piezo_model.prostor == False or self.kalibrace == False:
                    print(f"[{self.__class__.__name__}] FUNKCE PREDCASNE UKONCENA !!! -- NESPLNENA/Y PODMINKA")
                    self.controller.kalibrace_finish = False
                    break
                
                while self.mcu_model.lock_frekvence == False:
                    time.sleep(0.1) 
                    
                if iterace > 0:
                    #zapsat mcu frekvence a piezo polohu
                    self.vzorky = self.mcu_model.frekvence_vzorky.copy()
                    self.teplota = self.mcu_model.teplota_vzorky.copy()
                    self.tlak = self.mcu_model.tlak_vzorky.copy()
                    self.vlhkost = self.mcu_model.vlhkost_vzorky.copy()
                    self.osvetleni = self.mcu_model.osvetleni_vzorky.copy()
                    self.poloha = round(float(self.piezo_model.y_ref) * -1, 3)
                    if abs(self.poloha) < 1e-9:
                        self.poloha = 0.000
                    cas = datetime.now().strftime("%H:%M:%S.%f")[:-3]
                                    
                    df = pd.DataFrame({
                            "cas": [cas]*len(self.vzorky),
                            "pozice": [f"{self.poloha:.3f}"]*len(self.vzorky),
                            "frekvence": self.vzorky,
                            "teplota" : self.teplota,
                            "tlak": self.tlak,
                            "vlhkost": self.vlhkost,
                            "osvetleni" : self.osvetleni,
                            "smer" : [smer]*len(self.vzorky)
                        })
                    self.controller.zpracovani.df = pd.concat([self.controller.zpracovani.df, df], ignore_index = True)
                    
                    #ulozeni surovych dat do csv
                    df.to_csv(cesta_csv, mode='a', header=False, index=False)
                    
                    #zapis statistickych hodnot do dat a poslani do Zpracovani_modelu tridy
                    summary_df = pd.DataFrame([{
                        "cas": cas,
                        "pozice": self.poloha,
                        "frekvence_prumer": df["frekvence"].mean(),
                        "frekvence_std": df["frekvence"].std(),
                        "teplota_prumer": df["teplota"].mean(),
                        "tlak_prumer": df["tlak"].mean(),
                        "vlhkost_prumer": df["vlhkost"].mean(),
                        "osvetleni_prumer" : df["osvetleni"].mean(),
                        "smer" : smer
                    }])
                    self.controller.zpracovani.summary_df = pd.concat([self.controller.zpracovani.summary_df, summary_df], ignore_index=True)                
                                          
                    #pridani jednotlivych polozek vzorku do fronty
                    for f, t in zip(self.vzorky, self.teplota):
                        self.queue_graf.put({
                            "cas" : cas,
                            "pozice" : self.poloha,
                            "frekvence" : f,
                            "teplota" : t,
                            "smer" : smer
                        })
                    
                    print(f"[{self.__class__.__name__}] zápis dat do fronty!") 
  
                    if iterace < int(self.pocet_kroku):    
                        self.controller.M_C_nastav_pohyb_piezo(float(self.delka_kroku))
                        self.controller.M_C_pohyb_piezo(smer)
                        while not self.controller.lock_pohyb:
                            time.sleep(0.1)
                    
                    elif iterace == int(self.pocet_kroku):
                        self.controller.M_C_nastav_pohyb_piezo(float(self.delka_kroku))
                        self.controller.M_C_pohyb_piezo_GT(y=10000)
                        while not self.controller.lock_pohyb:
                            time.sleep(0.1)
                                       
                self.mcu_model.lock_frekvence = False
                
                if iterace <= int(self.pocet_kroku):
                    print(f"[{self.__class__.__name__}] iterace")
                    print(f"[{self.__class__.__name__}] CTENI FREKVENCE !!")
                    self.mcu_model.precist_frekvenci(int(self.pocet_zaznamu))
                
                iterace += 1
                
                #pokud vsechny pozice bylo zmereno, tak uspesna kalibrace
                if iterace == (self.pocet_kroku):
                    self.controller.kalibrace_finish = True
                                 
            #precteni celeho CSV a nahrani do XLSX
            df_csv = pd.read_csv(cesta_csv)
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Data"
                #hlavicka
                hlavicka = ["Čas (hh:mm:ss)", "Pozice (µm)", "Frekvence (Hz)", "Teplota (°C)", "Tlak (Pa)", "Vlhkost (%)", "Osvetleni (lux)", "Smer (y-/y+)"]
                bold_font = Font(bold=True)
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                #zapsat hlavicku
                for col_num, text in enumerate(hlavicka, start=1):
                    cell = ws.cell(row=1, column=col_num, value=text)
                    cell.font = bold_font
                #zapsat data
                for row_num, row_data in enumerate(df_csv.values, start=2):
                    for col_num, value in enumerate(row_data, start=1):
                        ws.cell(row=row_num, column=col_num, value=value)
                #ulozit
                wb.save(cesta_xlsx)
                print(f"[{self.__class__.__name__}] EXCEL VZORKY VYTVORENY, EXCEL (temp soubor) VYTVOREN")
                self.data_load(zpracovani_dat="pulzy", strategie_kalibrace="hystereze", delka_kroku=self.delka_kroku,merena_vzdalenost=self.merena_vzdalenost ,pocet_vzorku=self.pocet_zaznamu, pocet_kroku=self.pocet_kroku) 
                if self.controller.kalibrace_finish == False:
                    InfoMsg = f"Data byly uloženy ale kalibrace neproběhla celá"
                else:
                    InfoMsg = f"Kalibrace proběhla úspěšně"
                messagebox.showinfo("Kalibrace", InfoMsg)
            except Exception as e:
                self.kalibrace = False
                InfoMsg = f"CHYBA\nEXCEL (temp soubor) VZORKY NEVYTVORENY KONEC APLIKACE !!CHYBA!!\npravdepodobne otevreny soubor temp:\n{e}!!"
                messagebox.showinfo("Chyba", InfoMsg)
                print(f"[{self.__class__.__name__}] EXCEL VZORKY NEVYTVORENY --CHYBA!! {e}")   
                  
            print(f"[{self.__class__.__name__}] sběr dat hotový")  
            self.kalibrace = False #konec kalibrace
            self.mcu_model.lock_frekvence = True #zase odemknout pro pripad dalsiho mereni
            
            #debug vypis dat rychly
            display(self.controller.zpracovani.df)
            display(self.controller.zpracovani.summary_df)
                     
        self.t1 = threading.Thread(target=kalibrace_start_inner, daemon=True)
        self.t1.start()