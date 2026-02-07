import time
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from datetime import datetime
import shutil
from tkinter import messagebox
import win32com.client as win32

from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from controller.main_controller import MainController

class Zpracovani_model():
    def __init__(self, controller : "MainController"):
        self.controller = controller
        
        #Data z kalibrace
        self.kalibrace_df = pd.DataFrame()
        self.df = pd.DataFrame()
        self.summary_df = pd.DataFrame()
    
        #Data excel
        #Razitko #spolecne pro vsechny datasheety
        self.nazev = None
        self.nazev_pozice = "K52"
        self.katedra = None
        self.katedra_pozice = "A48"
        self.technicka_reference = None
        self.technicka_reference_pozice = "K48"
        self.kalibroval = None
        self.kalibroval_pozice = "R48"
        self.schvalil = None
        self.schvalil_pozice = "AB48"
        self.projekt = None
        self.projekt_pozice = "AL48"
        self.typ_dokumentu = None
        self.typ_dokumentu_pozice = "K50"
        self.status_dokumentu = None
        self.status_dokumentu_pozice = "AF50"
        self.cislo_dokumentu = None
        self.cislo_dokumentu_pozice = "AF52"
        self.univerzita = None
        self.univerzita_pozice = "A54"
        self.revize = None
        self.revize_pozice = "AF54"
        self.datum = None
        self.datum_pozice = "AJ54"
        self.jazyk = "cs"
        self.jazyk_pozice = "AS54"
        self.pocet_stran = None
        self.pocet_stran_pozice = "AW54"
    
        #MD - Main Datasheet
        #MD001 KALIBRACNI LIST NAMERENYCH HODNOT
        
        #Infromace o kalibraci
        self.typ_snimace = None
        self.typ_snimace_pozice = "O5"
        self.zpracovani_dat = None
        self.zpracovani_dat_pozice = "O6"
        self.strategie_kalibrace = None
        self.strategie_kalibrace_pozice = "O7"
        
        self.snimany_objekt = None
        self.snimany_objekt_pozice = "O10"
        self.snimany_material = None
        self.snimany_material_pozice = "O11"
        self.obvod_zpracovani = None
        self.obvod_zpracovani_pozice = "O12"
        self.napajeni_snimace = None
        self.napajeni_snimace_pozice = "O13"
        
        self.rozsah_mereni = None
        self.rozsah_mereni_pozice = "AK5"
        self.rozliseni_mereni = None
        self.rozliseni_mereni_pozice = "AK6"
        self.pocet_kroku = None
        self.pocet_kroku_pozice = "AK7"
        self.pocet_vzorku = None
        self.pocet_vzorku_pozice = "AK8"
        
        #Okolni podminky mereni
        self.teplota = None
        self.teplota_pozice = "O17"
        self.tlak = None
        self.tlak_pozice = "O18"
        self.relativni_vlhkost = None
        self.relativni_vlhkost_pozice = "O19"
        self.osvetleni = None
        self.osvetleni_pozice = "O20"
        
        #Namerene hodnoty
        self.merena_velicina = None
        self.maximalni_prumerna_hodnota = None
        self.minimalni_prumerna_hodnota = None
        
    def prirazeni_hodnot(self):
        
        #Infromace o kalibraci
        self.zpracovani_dat = str(self.kalibrace_df["zpracovani_dat"].iloc[0])
        self.strategie_kalibrace = str(self.kalibrace_df["strategie_kalibrace"].iloc[0])
        self.rozsah_mereni = str(self.kalibrace_df["rozsah_mereni"].iloc[0])
        self.rozliseni_mereni = str(self.kalibrace_df["rozliseni_mereni"].iloc[0])
        self.pocet_kroku = str(self.kalibrace_df["pocet_kroku"].iloc[0])
        self.pocet_vzorku = str(self.kalibrace_df["pocet_vzorku_na_krok"].iloc[0])
        
        #Okolni podminky mereni
        self.teplota = round(float(self.summary_df["teplota_prumer"].mean()), 1)
        self.tlak = round(float(self.summary_df["tlak_prumer"].mean()), 1)
        self.relativni_vlhkost = round(float(self.summary_df["vlhkost_prumer"].mean()), 1)
        self.osvetleni = round(float(self.summary_df["osvetleni_prumer"].mean()), 1)
        
    def vytvorit_excel(self):
        #nacteni template - ORIGINALNI TEMPLATE - ZKOPIROVANI
        try:
            shutil.copy("template/Datasheet_senzor_template.xlsx", "vyhodnoceni.xlsx")
        except PermissionError:
            InfoMsg = f"[{self.__class__.__name__}] EXCEL SOUBOR JE JIŽ OTEVŘENÝ"
            messagebox.showerror("Excel", InfoMsg)
            print(InfoMsg)
            return
        #KOPIE POD JINYM NAZVEM 
        wb = load_workbook("vyhodnoceni.xlsx")
        #VYBRANY LIST AKTUALNI - mozna do controlleru
        
        #prirazeni do excelu - razitko - vsude stejne
        for sheet in wb.worksheets:
            if sheet.title == "MD005":
                continue
            sheet[self.nazev_pozice] = self.nazev
            sheet[self.katedra_pozice] = self.katedra
            sheet[self.technicka_reference_pozice] = self.technicka_reference
            sheet[self.kalibroval_pozice] = self.kalibroval
            sheet[self.schvalil_pozice] = self.schvalil
            sheet[self.projekt_pozice] = self.projekt
            sheet[self.status_dokumentu_pozice] = self.status_dokumentu
            sheet[self.cislo_dokumentu_pozice] = self.cislo_dokumentu
            sheet[self.univerzita_pozice] = self.univerzita
            sheet[self.revize_pozice] = self.revize
            sheet[self.datum_pozice] = self.datum
            sheet[self.jazyk_pozice] = self.jazyk
        
        #navrat to prvniho sheetu MD001
        self.sheet = wb["MD001"]
        
        #priprazeni do excelu - informace o kalibraci
        self.sheet[self.typ_snimace_pozice] = str(self.typ_snimace)
        self.sheet[self.zpracovani_dat_pozice] = str(self.zpracovani_dat)
        self.sheet[self.strategie_kalibrace_pozice] = str(self.strategie_kalibrace)
        self.sheet[self.snimany_objekt_pozice] = str(self.snimany_objekt)
        self.sheet[self.snimany_material_pozice] = str(self.snimany_material)
        self.sheet[self.obvod_zpracovani_pozice] = str(self.obvod_zpracovani)
        self.sheet[self.napajeni_snimace_pozice] = str(self.napajeni_snimace)
        self.sheet[self.rozsah_mereni_pozice] = str(self.rozsah_mereni)
        self.sheet[self.rozliseni_mereni_pozice] = str(self.rozliseni_mereni)  
        self.sheet[self.pocet_kroku_pozice] = str(self.pocet_kroku)
        self.sheet[self.pocet_vzorku_pozice] = str(self.pocet_vzorku)
        
        #priprazeni do excelu - okolni podminky mereni
        self.sheet[self.teplota_pozice] = str(self.teplota)
        self.sheet[self.tlak_pozice] = str(self.tlak)
        self.sheet[self.relativni_vlhkost_pozice] = str(self.relativni_vlhkost)
        self.sheet[self.osvetleni_pozice] = str(self.osvetleni)
    
        #pridani dat z kalibrace do MD005
        pracovni_slozka = os.path.join(self.controller.kalibrace.pracovni_slozka, "temp.xlsx")
        wb_data = load_workbook(pracovni_slozka)
        ws_data = wb_data["Data"]
        self.sheet = wb["MD005"]
        
        for radek in ws_data.iter_rows(values_only=True):
            print(radek)
            self.sheet.append(radek)
        
        #vytvoreni grafu prubehu merenych hodnot
        self.sheet = wb["MD002"]
        # img_cesta = os.path.join(self.controller.kalibrace.pracovni_slozka, "graf_frekvence.png") #TODO OPRAVIT JMENO SOUBORU OBRAZKU !!!
        img_cesta = self.controller.kalibrace.kalibracni_obrazek
        img = Image(img_cesta)
        img.width = 573
        img.height = 427
        img.anchor = "C7"
        self.sheet.add_image(img)
        
        #ulozeni do pracovni slozky
        datum = datetime.now().strftime("%Y-%m-%d")
        poradi = 0
        jmeno_souboru = f"vyhodnoceni_{datum}_{poradi:03}.xlsx"
        je_v_adresari = True
        while(je_v_adresari is True):
            cesta = os.path.join(self.controller.kalibrace.pracovni_slozka, jmeno_souboru)
            if os.path.exists(cesta):
                je_v_adresari = True
                poradi += 1
                jmeno_souboru = f"vyhodnoceni_{datum}_{poradi:03}.xlsx"
            else:
                je_v_adresari = False
        
        pracovni_slozka = os.path.join(self.controller.kalibrace.pracovni_slozka, jmeno_souboru)
        wb.save(pracovni_slozka)
        wb.close()
        del wb
        time.sleep(2)
        
        jmeno_pdf = f"vyhodnoceni_{datum}_{poradi:03}.pdf"
        pracovni_slozka_pdf = os.path.join(self.controller.kalibrace.pracovni_slozka, jmeno_pdf)
        
        print(f"[{self.__class__.__name__}] VYTVORENY .xlsx a .pdf SOUBORY s nazvy {jmeno_souboru}")
        
        # self.vytvorit_pdf(pracovni_slozka, pracovni_slozka_pdf, ["MD001", "MD002", "MD003", "MD004"])
        
        #otevreni souboru
        os.startfile(pracovni_slozka)
        if os.path.exists(pracovni_slozka_pdf):
           os.startfile(pracovni_slozka_pdf)
        else:
           print(f"[{self.__class__.__name__}]PDF soubor nebyl nalezen!")
        
    def vytvorit_pdf(self, xlsx_cesta, pdf_cesta, sheets):
        if os.path.exists(pdf_cesta):
            os.remove(pdf_cesta)
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        wb = excel.Workbooks.Open(xlsx_cesta)
        
        excel.Sheets(sheets).Select()
        # wb.Save()
        
        print(f"[{self.__class__.__name__} DEBUG]")
        time.sleep(5)
        
        excel.ActiveSheet.ExportAsFixedFormat(0, pdf_cesta)
        
        time.sleep(5)
        wb.Close(False)
        excel.Quit()